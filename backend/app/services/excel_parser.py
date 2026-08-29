"""Excel (.xlsx) parser (Phase 3: upload pipeline).

Reads the first worksheet, validates required columns and returns the same
standardised row dictionaries as csv_parser. Corrupted files are reported
as FileParseError so the batch can be marked Failed.
"""
import io
import zipfile
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backend.app.services.parser_service import (
    EmptyFileError,
    FileParseError,
    MissingColumnsError,
    REQUIRED_COLUMNS,
    resolve_header,
)


def _cell_to_text(value) -> str:
    """Normalise a single cell value into its transport string form."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _extract_rows(rows):
    """Return (header_row, data_rows): the first non-empty row is the header."""
    for index, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() for cell in row):
            return row, rows[index + 1:]
    return None, []


def _row_dict(row, headers):
    return {header: _cell_to_text(value) for header, value in zip(headers, row)}


def parse_excel(content: bytes) -> list:
    if not content or not content.strip():
        raise EmptyFileError("The uploaded file is empty.")

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except (zipfile.BadZipFile, InvalidFileException, ValueError, OSError) as exc:
        raise FileParseError("Corrupted or invalid Excel file.") from exc

    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_row, data_rows = _extract_rows(rows)
    if header_row is None:
        raise EmptyFileError("The Excel file has no non-empty worksheet (no header row found).")

    headers = [resolve_header(header) for header in header_row]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise MissingColumnsError(
            f"Missing required columns: {', '.join(sorted(missing))}. "
            f"Found: {[h for h in headers if _cell_to_text(h)]}."
        )

    parsed = []
    for row in data_rows:
        record = _row_dict(row, headers)
        if any(record.values()):
            parsed.append(record)

    if not parsed:
        raise EmptyFileError("The Excel file contains no data rows.")
    return parsed